import os
from typing import Type
import openpyxl
from openpyxl import load_workbook



# def Bet_Analysis(xcl_file):
#     wb = openpyxl.load_workbook(xcl_file)
#     sheet = wb.active


#     race_info = os.path.basename(xcl_file)
#     race_info = os.path.splitext(race_info)[0]
#     print(race_info)

#     bf_list = []
#     bk_list = []


#     last_cell = len(sheet["A"])
#     count = 3

#     # Bf list
#     while count <= last_cell:
#         sheet_pos = "A" + str(count)
#         x = sheet[sheet_pos].value
#         if x == "EW" or x == "MD" or x == "SUS" or x == None or x is None:
#             x = 1000
#         else:
#             x = float(x)
#         bf_list.append(x)
#         count += 1

#     count = 3
#     # Bk list
#     while count <= last_cell:
#         sheet_pos = "B" + str(count)
#         x = sheet[sheet_pos].value
#         if x == "EW" or x == "MD" or x == "SUS" or x == None or x is None:
#             x = 1
#         else:
#             x = float(x)
#         bk_list.append(x)
#         count += 1


#     #calculations
#     if len(bf_list) != 0:
#         if None in bf_list or None in bk_list:
#             pass
#         else:
#             list_index = 0
#             while list_index < len(bf_list):
#                 bf = float(bf_list[list_index])
#                 bk = float(bk_list[list_index])
#                 if bk >= 20 and bf <= 150:
#                     temp_bets = openpyxl.load_workbook("TempBetsPlaced.xlsx")
#                     tb_sheet = temp_bets.active
#                     tb_row = len(tb_sheet["A"]) + 1
#                     placement = list_index + 1

#                     tb_sheet_pos_info = "A" + str(tb_row)
#                     tb_sheet[tb_sheet_pos_info] = race_info
#                     tb_sheet_bet_info = "B" + str(tb_row)
#                     tb_sheet[tb_sheet_bet_info] = placement
#                     tb_sheet_odd_info = "C" + str(tb_row)
#                     tb_sheet[tb_sheet_odd_info] = bk

#                     temp_bets.save("TempBetsPlaced.xlsx")


                    
                    
#                 list_index += 1



# Version 1
def Bet_Analysis(xcl_file):
    wb = openpyxl.load_workbook(xcl_file)
    sheet = wb.active


    race_info = os.path.basename(xcl_file)
    race_info = os.path.splitext(race_info)[0]
    print(race_info)

    bf_list = []
    bk_list = []


    last_cell = len(sheet["A"])
    count = 3

    # Bf list
    while count <= last_cell:
        sheet_pos = "A" + str(count)
        x = sheet[sheet_pos].value
        if x == "EW" or x == "MD" or x == "SUS" or x == None or x is None:
            x = 1000
        else:
            x = float(x)
        bf_list.append(x)
        count += 1

    count = 3
    # Bk list
    while count <= last_cell:
        sheet_pos = "B" + str(count)
        x = sheet[sheet_pos].value
        if x == "EW" or x == "MD" or x == "SUS" or x == None or x is None:
            x = 1
        else:
            x = float(x)
        bk_list.append(x)
        count += 1


    print(bf_list)
    print(bk_list)

    #calculations
    if len(bf_list) != 0:
        if None in bf_list or None in bk_list:
            pass
        else:
            list_index = 0
            while list_index < len(bf_list):
                bf = float(bf_list[list_index])
                bk = float(bk_list[list_index])
                #t = bk + bk * 0.2
                #margin = bk/bf
                #if margin >= 0.9 and margin <= 1.5:
                if bf <= 150 and bk >= 30: 
                    #if bk < 100:
                    temp_bets = openpyxl.load_workbook("TempBetsPlaced.xlsx")
                    tb_sheet = temp_bets.active
                    tb_row = len(tb_sheet["A"]) + 1
                    placement = list_index + 1

                    tb_sheet_pos_info = "A" + str(tb_row)
                    tb_sheet[tb_sheet_pos_info] = race_info
                    tb_sheet_bet_info = "B" + str(tb_row)
                    tb_sheet[tb_sheet_bet_info] = placement
                    tb_sheet_odd_info = "C" + str(tb_row)
                    tb_sheet[tb_sheet_odd_info] = bk

                    temp_bets.save("TempBetsPlaced.xlsx")

                if bf <= 50 and bk >= 20 and bk < 30: 
                    #if bk < 100:
                    temp_bets = openpyxl.load_workbook("TempBetsPlaced.xlsx")
                    tb_sheet = temp_bets.active
                    tb_row = len(tb_sheet["A"]) + 1
                    placement = list_index + 1

                    tb_sheet_pos_info = "A" + str(tb_row)
                    tb_sheet[tb_sheet_pos_info] = race_info
                    tb_sheet_bet_info = "B" + str(tb_row)
                    tb_sheet[tb_sheet_bet_info] = placement
                    tb_sheet_odd_info = "C" + str(tb_row)
                    tb_sheet[tb_sheet_odd_info] = bk

                    temp_bets.save("TempBetsPlaced.xlsx")

                if bf <= 35 and bk >= 10 and bk < 20: 
                    #if bk < 100:
                    temp_bets = openpyxl.load_workbook("TempBetsPlaced.xlsx")
                    tb_sheet = temp_bets.active
                    tb_row = len(tb_sheet["A"]) + 1
                    placement = list_index + 1

                    tb_sheet_pos_info = "A" + str(tb_row)
                    tb_sheet[tb_sheet_pos_info] = race_info
                    tb_sheet_bet_info = "B" + str(tb_row)
                    tb_sheet[tb_sheet_bet_info] = placement
                    tb_sheet_odd_info = "C" + str(tb_row)
                    tb_sheet[tb_sheet_odd_info] = bk

                    temp_bets.save("TempBetsPlaced.xlsx")


                # if margin >= 0.95 and margin <= 1.5:
                #     if bk >= 13 and bk < 28:
                #         temp_bets = openpyxl.load_workbook("TempBetsPlaced.xlsx")
                #         tb_sheet = temp_bets.active
                #         tb_row = len(tb_sheet["A"])
                #         placement = list_index + 1

                #         tb_sheet_pos_info = "A" + str(tb_row)
                #         tb_sheet[tb_sheet_pos_info] = race_info
                #         tb_sheet_bet_info = "B" + str(tb_row)
                #         tb_sheet[tb_sheet_bet_info] = placement
                #         tb_sheet_odd_info = "C" + str(tb_row)
                #         tb_sheet[tb_sheet_odd_info] = bk

                #         temp_bets.save("TempBetsPlaced.xlsx")

                list_index += 1





directory = (r"D:\Python_Repos\Excel_Scraper\new week")

files_list = []

for filename in os.listdir(directory):
    if filename.endswith("xlsx"):
        #print(os.path.join(directory, filename))
        files_list.append(os.path.join(directory, filename))
        continue
    else:
        continue

for b in files_list:
    Bet_Analysis(b)

